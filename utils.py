from collections import defaultdict
import pandas as pd
from io import BytesIO
import xlsxwriter
from openpyxl import load_workbook
from PIL import Image, ExifTags
import streamlit as st
import os


# environment variables
os.environ["ENV"] = st.secrets["aws"]["ENV"]


# Image utilities
def handle_image_orientation(image: Image.Image) -> Image.Image:
    """
    Handle image orientation based on EXIF data.
    
    Args:
        image: PIL Image object
        
    Returns:
        PIL Image object with correct orientation
    """
    try:
        for orientation in ExifTags.TAGS.keys():
            if ExifTags.TAGS[orientation] == 'Orientation':
                break
        exif = dict(image._getexif().items())

        if exif[orientation] == 3:
            image = image.rotate(180, expand=True)
        elif exif[orientation] == 6:
            image = image.rotate(270, expand=True)
        elif exif[orientation] == 8:
            image = image.rotate(90, expand=True)
    except (AttributeError, KeyError, IndexError):
        print("Image does not have exif data")
        # cases: image don't have exif data
        pass
    
    return image

# substrate analysis
def generate_keys(key_list, multiplier = 3):
    new_list = []

    for label in key_list:
        new_list.extend([label]*multiplier)

    return new_list


def create_substrate_dataframe(response_data: dict, csv_name: str) -> pd.DataFrame:
    segment_distances = ["0 - 19.5m", "25 - 44.5m", "50 - 65.5m", "75 - 94.5m"]

    # pop out the slate info 
    info_segment = response_data.pop('info_segment', None)
    # get unique keys
    response_keys = list(response_data.keys())
    # create dataframes
    df = pd.concat([pd.DataFrame.from_dict(response_data[key]) for key in list(response_data.keys())], axis = 1)

    # create unique column names
    column_names = []
    for num in range(len(list(response_data.keys()))):
        column_names.extend([f"distance_{num}", f"label_{num}", f"clear_{num}"])
    # set column names
    df.columns = column_names

    # generate multi-level index
    segment_list = generate_keys(list(response_data.keys()))
    merge_list = generate_keys(segment_distances)

    # create multi index
    arrays = [
        segment_list,
        merge_list,
        column_names
    ]
    columns = pd.MultiIndex.from_arrays(arrays)
    # create a dataframe
    df = pd.DataFrame(df.iloc[:,:].values, columns=columns)  

    # save the dataframe to a csv file
    df.to_csv(csv_name, index=False)

    return df, info_segment


def extract_details(info: dict) -> list:

    return [info["distance"], info["label"], info["label_status"]]


def extract_single_attributes(selected_set: list, index_val: int) -> list:
    sub_segments = []
    # first segment
    first_set = selected_set[index_val]
    second_set = selected_set[index_val + 20]
    sub_segments.extend(extract_details(first_set))
    sub_segments.extend(extract_details(second_set))

    return sub_segments


def substrate_excel_creation(response_data: dict, info_data: dict, excel_name: str):
    information_data = info_data
    selected_set_one = response_data["segment_one"]
    selected_set_two = response_data["segment_two"]
    selected_set_three = response_data["segment_three"]
    selected_set_four = response_data["segment_four"]


    final_segments = []
    for index in range(20):
        sub_set_segments = []
        # first segment
        sub_set_segments.extend(extract_single_attributes(selected_set_one, index))
        # seconds segment
        sub_set_segments.extend(extract_single_attributes(selected_set_two, index))
        # seconds segment
        sub_set_segments.extend(extract_single_attributes(selected_set_three, index))
        # seconds segment
        sub_set_segments.extend(extract_single_attributes(selected_set_four, index))
        # append the segment
        final_segments.append(sub_set_segments)

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(excel_name)
    worksheet = workbook.add_worksheet()

    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True, 'center_across': True, 'border': True})

    # Add a border
    border = workbook.add_format({'border': True})

    # Add a number format for cells with money.
    not_clear = workbook.add_format({'bold': True, 'bg_color': 'red', 'border': True})

    #Write info headers.
    worksheet.merge_range("A1:B1", "Site Name", bold)
    worksheet.merge_range("C1:D1", "Country/ island", bold)
    worksheet.merge_range("E1:F1", "Team Leader", bold)
    worksheet.merge_range("G1:H1", "Data Recorded By", bold)
    worksheet.merge_range("I1:J1", "Depth", bold)
    worksheet.merge_range("K1:L1", "Date", bold)
    worksheet.merge_range("M1:N1", "Time", bold)
    
    
    # Write some data headers.
    worksheet.merge_range("A5:P5", "Substrate Analysis", bold)
    worksheet.merge_range("A6:D6", "Segment One", bold)
    worksheet.merge_range("E6:H6", "Segment Two", bold)
    worksheet.merge_range("I6:L6", "Segment Three", bold)
    worksheet.merge_range("M6:P6", "Segment Four", bold)


    # distances
    worksheet.merge_range("A7:D7", "0 - 19.5m", bold)
    worksheet.merge_range("E7:H7", "25 - 44.5m", bold)
    worksheet.merge_range("I7:L7", "50 - 69.5m", bold)
    worksheet.merge_range("M7:P7", "75 - 94.5", bold)

    info_col = 0
    info_row = 2

    info_fields = [
        "site_name",
        "country_island",
        "team_leader",
        "data_recorded_by",
        "depth",
        "date",
        "time",
    ]

    for field in info_fields:
        worksheet.merge_range(
            info_row - 1,        # row index (0-based)
            info_col,
            info_row - 1,
            info_col + 1,
            info_data.get(field, ""),
            border
        )
        info_col += 2
            

    
    row = 8
    # adding records
    for diff_segments in final_segments:
        col = 0
        for range_index in range(0, 24, 3):
            worksheet.write(row, col, diff_segments[range_index], border)
            col += 1
            worksheet.write(row, col, diff_segments[range_index +1], not_clear if not diff_segments[range_index +2] else border)
            col += 1
        row += 1


    workbook.close()


def load_and_prepare_excel_for_substrate(excel_name: str):
    # Load the workbook and select the active sheet
    workbook = load_workbook(excel_name)
    with BytesIO() as buffer:
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def create_fish_slate_dataframe(response_data: dict, csv_name: str) -> pd.DataFrame:
    # distances are constants
    distances = ["0 - 20m", "25 - 45m", "50 - 75m", "75 - 95m"]

    main_keys = list(response_data.keys())

    info_df = pd.concat([pd.DataFrame.from_dict(response_data[key_]) for key_ in main_keys])
    
    new_columns = []
    new_columns.append("name")
    for distance_index in range(len(distances)):
        new_columns.extend([distances[distance_index], "set_{}_clear".format(distance_index)])

    info_df.columns = new_columns

    # save the dataframe to a csv file
    info_df.to_csv(csv_name, index=False)

    return info_df


def fish_slate_excel_creation(response_data: dict, info_data: dict, excel_name: str):

    data_list = []

    for key_ in list(response_data.keys()):
        data_list.extend(response_data[key_])


    distances = ["0 - 20m", "25 - 45m", "50 - 75m", "75 - 95m"]
    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(excel_name)
    worksheet = workbook.add_worksheet()

    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True, 'center_across': True, 'border': True})

    # sub category
    sub_format = workbook.add_format({'bold': True, 'center_across': True, 'border': True, 'bg_color': 'green'})

    # Add a border
    border = workbook.add_format({'border': True})

    # Add a number format for cells with money.
    not_clear = workbook.add_format({'bold': True, 'bg_color': 'red', 'border': True})

    #Write info headers.
    worksheet.merge_range("A1:B1", "Site Name", bold)
    worksheet.merge_range("C1:D1", "Country/ island", bold)
    worksheet.merge_range("E1:F1", "Team Leader", bold)
    worksheet.merge_range("G1:H1", "Data Recorded By", bold)
    worksheet.merge_range("I1:J1", "Depth", bold)
    worksheet.merge_range("K1:L1", "Date", bold)
    worksheet.merge_range("M1:N1", "Time", bold)
    
    
    
    worksheet.merge_range("A4:E4", "Fish Slate Analysis", bold)
    worksheet.write(4, 0, "Type", bold)
    worksheet.set_column('A:A', 30)
    
    # Write Fish Slate record data
    info_col = 0
    info_row = 2

    info_fields = [
        "site_name",
        "country_island",
        "team_leader",
        "data_recorded_by",
        "depth",
        "date",
        "time",
    ]

    for field in info_fields:
        worksheet.merge_range(
            info_row - 1,        # row index (0-based)
            info_col,
            info_row - 1,
            info_col + 1,
            info_data.get(field, ""),
            border
        )
        info_col += 2
    
    
    
    # distances
    for index in range(len(distances)):
        worksheet.write(4, index + 1, distances[index] , bold)

    row = 5
    col = 0
    for key_ in list(response_data.keys()):
        worksheet.write(row, col, key_, sub_format)
        row +=1 
        for data_dict in response_data[key_]:
            worksheet.write(row, col, data_dict["name"], border)
            worksheet.write(row, col+1, data_dict["distance_one"], not_clear if not data_dict["distance_one_clear"] else border)
            worksheet.write(row, col+2, data_dict["distance_two"], not_clear if not data_dict["distance_two_clear"] else border)
            worksheet.write(row, col+3, data_dict["distance_three"], not_clear if not data_dict["distance_three_clear"] else border)
            worksheet.write(row, col+4, data_dict["distance_four"], not_clear if not data_dict["distance_four_clear"] else border)
            row += 1


    workbook.close()


def load_and_prepare_excel_for_fish_slate(excel_name: str):
    # Load the workbook and select the active sheet
    workbook = load_workbook(excel_name)
    with BytesIO() as buffer:
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def substrate_excel_data_extractor(data: pd.DataFrame) -> dict:
    suffixes = ["one", "two", "three", "four"]
    annots = defaultdict(list)
    column_list = list(data.columns)
    count = 0

    for index in range(0, 12, 3):
        distance = data[column_list[index]].to_list()
        label = data[column_list[index + 1]].to_list()
        status = data[column_list[index + 2]].to_list()

        for distance_, label_, status_ in zip(distance, label, status):
            annots["segment_{}".format(suffixes[count])].append({
                "distance": distance_,
                "label": label_,
                "label_status": status_
            })

        count +=1

    return dict(annots)


def extract_fish_details(fish_details: list) -> list:
    total_records = []

    for record_ in fish_details:
        sample_dict = {}
        sample_dict["name"] = record_["name"]
        sample_dict["distance_one"] = record_["0 - 20m"]
        sample_dict["distance_one_clear"] = record_["set_0_clear"]
        sample_dict["distance_two"] = record_["25 - 45m"]
        sample_dict["distance_two_clear"] = record_["set_1_clear"]
        sample_dict["distance_three"] = record_["50 - 75m"]
        sample_dict["distance_three_clear"] = record_["set_2_clear"]
        sample_dict["distance_four"] = record_["75 - 95m"]
        sample_dict["distance_four_clear"] = record_["set_3_clear"]
        total_records.append(sample_dict)

    return total_records


def fish_excel_data_extractor(data: pd.DataFrame) -> dict:
    # get data records
    records = data.to_dict(orient='records')

    annots = defaultdict(list)

    # original records
    fish_records = records[: 12]
    invertebrates_records = records[12: 26]
    impacts_records = records[26: 33]
    coral_disease_records = records[33: 35]
    rare_animals_records = records[35: 39]

    # processed records
    process_dict = {
        "fish": fish_records,
        "invertebrates": invertebrates_records,
        "impacts": impacts_records,
        "coral_disease": coral_disease_records,
        "rare_animals": rare_animals_records

    }
    
    for key_ in process_dict.keys():
        annots[key_].extend(extract_fish_details(process_dict[key_]))

    return dict(annots)